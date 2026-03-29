"""
Import Maestro de Items from SIESA Excel into normalized Supabase schema.

Tables populated:
  1. classifications  — dimension value normalization (categoria, subcategoria, etc.)
  2. warehouses       — normalized bodegas
  3. products         — base products (one per reference base)
  4. product_variants — each unique SIESA reference (acabado, aleacion, length variants)
  5. product_warehouse — per-variant per-bodega financials, dates, ABC
  6. import_history   — audit log
"""
import os
import sys
import re
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import openpyxl
from supabase import create_client

# ── Config ──
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://cnuhbxtwezrhwpmdubfo.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# Subcategorias that are profiles (have length in reference)
PROFILE_SUBCATEGORIAS = {"PERFILES DE ALUMINIO", "PERFILES LIGHT"}

# ── Column mapping: Excel header -> our field name ──
COLUMN_MAP = {
    "Item": "item_id",
    "Desc. bodega": "bodega",
    "Referencia": "reference",
    "Estado": "status",
    "CATEGORIA": "categoria_raw",
    "SISTEMAS": "sistema_raw",
    "SUBCATEGORIA": "subcategoria_raw",
    "LINEA": "linea_raw",
    "Desc. ext. 2 detalle": "acabado_desc_raw",
    "Desc. item": "description",
    "Ext. 1 detalle": "aleacion_raw",
    "Ext. 2 detalle": "acabado_code_raw",
    "Compra": "flag_compra",
    "Venta": "flag_venta",
    "Manufactura": "flag_manufactura",
    "Maneja lote": "flag_lote",
    "Maneja paquete": "flag_paquete",
    "Costo prom. uni.": "costo_promedio",
    "Precio unit.": "precio_unitario",
    "Costo prom. total": "costo_promedio_total",
    "Peso U.M. Inv.": "peso_um",
    "Unidad de precio": "unidad_precio",
    "abc rotac. veces item": "abc_rotacion_veces",
    "abc rotac. costo item": "abc_rotacion_costo",
    "abc rotac. veces bod.": "abc_rotacion_veces_bod",
    "abc rotac. costo bod.": "abc_rotacion_costo_bod",
    "Fecha ultima venta": "fecha_ultima_venta",
    "Fecha ultima salida": "fecha_ultima_salida",
    "Fecha ultima entrada": "fecha_ultima_entrada",
    "Fecha ultima compra": "fecha_ultima_compra",
    "Fecha ultimo conteo": "fecha_ultimo_conteo",
    "Fecha creacion": "fecha_creacion_siesa",
    "Fecha actualizacion": "fecha_actualizacion_siesa",
    "Fecha inactivacion": "fecha_inactivacion",
    "Desc. posicion arancelaria": "posicion_arancelaria",
    "Margen(%)": "margen_pct",
}

# Build normalized key map (strip accents for matching)
COLUMN_MAP_NORMALIZED = {}
for k, v in COLUMN_MAP.items():
    COLUMN_MAP_NORMALIZED[k] = v
    stripped = k.replace("a\u0301", "a").replace("e\u0301", "e").replace("i\u0301", "i").replace("o\u0301", "o").replace("u\u0301", "u")
    stripped = stripped.replace("\u00e1", "a").replace("\u00e9", "e").replace("\u00ed", "i").replace("\u00f3", "o").replace("\u00fa", "u")
    COLUMN_MAP_NORMALIZED[stripped] = v


# ═══════════════════════════════════════════════════════
# Helper functions
# ═══════════════════════════════════════════════════════

def normalize_classification(dimension, raw_value):
    if not raw_value:
        return None
    raw = str(raw_value).strip()
    if not raw:
        return None

    code = None
    name = raw
    match = re.match(r'^(\d+)\s*-\s*(.+)$', raw)
    if match:
        code = match.group(1).strip()
        name = match.group(2).strip()

    name = re.sub(r'\s+', ' ', name).strip()
    if dimension in ('categoria', 'linea', 'sistema'):
        name = name.title()
    elif dimension == 'subcategoria':
        name = name.title()
        for prep in [' De ', ' Del ', ' La ', ' Las ', ' Los ', ' En ', ' No ']:
            name = name.replace(prep, prep.lower())

    return {
        "dimension": dimension,
        "original_value": raw,
        "normalized_value": name,
        "code": code,
        "status": "active",
        "created_by": "import_maestro",
    }


def parse_bool(val):
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ("si", "s\u00ed", "yes", "true", "1", "s")


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    try:
        return datetime.strptime(str(val), "%Y-%m-%d").isoformat()
    except (ValueError, TypeError):
        return None


def parse_number(val, default=0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def extract_base_ref_and_length(reference, subcategoria_raw):
    """Split a SIESA reference into base_ref and length_m.

    For profiles: "STM-297 x 6.00" -> ("STM-297", 6.00)
    For non-profiles: "C030110" -> ("C030110", None)
    """
    is_profile = subcategoria_raw and subcategoria_raw.strip().upper() in PROFILE_SUBCATEGORIAS

    if is_profile and ' x ' in reference:
        parts = reference.rsplit(' x ', 1)
        base = parts[0].strip()
        length_str = parts[1].strip()
        # Extract numeric part (handle "6.00", "6 NO USAR", etc.)
        length_match = re.match(r'^([\d.]+)', length_str)
        length_m = float(length_match.group(1)) if length_match else None
        return base, length_m

    return reference, None


def classify_warehouse(name):
    """Extract city and type from warehouse name."""
    name_upper = name.upper().strip()
    wtype = "bodega"
    if "AVERIAS" in name_upper or "AVERIA" in name_upper:
        wtype = "averias"
    elif "CONTENEDOR" in name_upper:
        wtype = "contenedores"

    # Extract city: last word usually
    city = None
    for city_name in ["TENJO", "PALMIRA", "BARRANQUILLA", "MUNDIAL"]:
        if city_name in name_upper:
            city = city_name.title()
            break

    return city, wtype


# ═══════════════════════════════════════════════════════
# Excel reading and column mapping
# ═══════════════════════════════════════════════════════

def read_excel(filepath):
    print(f"Reading: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(row)]
            continue
        rows.append(row)
    wb.close()
    print(f"  {len(rows)} rows, {len(headers)} columns")
    return headers, rows


def map_columns(headers):
    col_index = {}
    for i, header in enumerate(headers):
        if header in COLUMN_MAP_NORMALIZED:
            col_index[COLUMN_MAP_NORMALIZED[header]] = i
            continue
        clean = header.replace("\ufffd", "a").replace("\u00f3", "o").replace("\u00fa", "u").replace("\u00e9", "e").replace("\u00e1", "a")
        if clean in COLUMN_MAP_NORMALIZED:
            col_index[COLUMN_MAP_NORMALIZED[clean]] = i
            continue
        for key, field in COLUMN_MAP_NORMALIZED.items():
            if key.lower() in header.lower() or header.lower() in key.lower():
                if field not in col_index:
                    col_index[field] = i
                    break
    print(f"  Mapped {len(col_index)}/{len(COLUMN_MAP)} columns")
    unmapped = set(COLUMN_MAP.values()) - set(col_index.keys())
    if unmapped:
        print(f"  WARNING: Unmapped fields: {unmapped}")
    return col_index


# ═══════════════════════════════════════════════════════
# Step 1: Classifications (same as before)
# ═══════════════════════════════════════════════════════

def extract_and_upsert_classifications(rows, col_index):
    dims = {d: set() for d in ("categoria", "subcategoria", "sistema", "acabado", "linea", "aleacion")}
    field_to_dim = {
        "categoria_raw": "categoria", "subcategoria_raw": "subcategoria",
        "sistema_raw": "sistema", "acabado_desc_raw": "acabado",
        "linea_raw": "linea", "aleacion_raw": "aleacion",
    }
    for row in rows:
        for field, dim in field_to_dim.items():
            if field in col_index:
                val = row[col_index[field]]
                if val and str(val).strip():
                    dims[dim].add(str(val).strip())

    total = sum(len(v) for v in dims.values())
    print(f"\nClassifications found: {total}")
    for dim, values in dims.items():
        print(f"  {dim}: {len(values)}")

    records = []
    for dim, values in dims.items():
        for raw in sorted(values):
            rec = normalize_classification(dim, raw)
            if rec:
                records.append(rec)

    for i in range(0, len(records), 200):
        batch = records[i:i + 200]
        supabase.table("classifications").upsert(batch, on_conflict="dimension,original_value").execute()

    print(f"  {len(records)} classifications upserted")
    return len(records)


# ═══════════════════════════════════════════════════════
# Step 2: Warehouses
# ═══════════════════════════════════════════════════════

def extract_and_upsert_warehouses(rows, col_index):
    print(f"\nExtracting warehouses...")
    bodega_names = set()
    for row in rows:
        if "bodega" in col_index:
            val = row[col_index["bodega"]]
            if val and str(val).strip():
                bodega_names.add(str(val).strip())

    records = []
    for name in sorted(bodega_names):
        city, wtype = classify_warehouse(name)
        records.append({"name": name, "city": city, "type": wtype})

    supabase.table("warehouses").upsert(records, on_conflict="name").execute()
    print(f"  {len(records)} warehouses upserted")

    # Fetch back to get id mapping
    result = supabase.table("warehouses").select("id, name").execute()
    warehouse_map = {r["name"]: r["id"] for r in result.data}
    return warehouse_map


# ═══════════════════════════════════════════════════════
# Step 3: Products + Variants + Warehouse junction
# ═══════════════════════════════════════════════════════

def import_products_and_variants(rows, col_index, warehouse_map, batch_id):
    print(f"\nProcessing products, variants, and warehouse data...")

    def get(row, field, default=None):
        if field in col_index:
            val = row[col_index[field]]
            if val is not None:
                return val
        return default

    # ── Group all rows by full SIESA reference ──
    # Each group = same reference across multiple bodegas
    ref_groups = defaultdict(list)
    skipped = 0
    for row in rows:
        ref = get(row, "reference")
        if not ref or not str(ref).strip():
            skipped += 1
            continue
        ref_groups[str(ref).strip()].append(row)

    print(f"  {len(rows)} rows -> {len(ref_groups)} unique SIESA references ({skipped} skipped)")

    # ── Build base products and variants ──
    # base_ref -> product data
    product_data = {}
    # full_ref -> variant data
    variant_data = {}
    # (full_ref, warehouse_id) -> warehouse row data (dict for dedup)
    pw_data = {}

    for full_ref, group in ref_groups.items():
        first_row = group[0]
        subcategoria = str(get(first_row, "subcategoria_raw", "")).strip()
        base_ref, length_m = extract_base_ref_and_length(full_ref, subcategoria)
        is_profile = subcategoria.upper() in PROFILE_SUBCATEGORIAS

        # ── Build/update base product ──
        if base_ref not in product_data:
            product_data[base_ref] = {
                "reference": base_ref,
                "description": str(get(first_row, "description", "")).strip() or None,
                "categoria_raw": str(get(first_row, "categoria_raw", "")).strip() or None,
                "subcategoria_raw": subcategoria or None,
                "sistema_raw": str(get(first_row, "sistema_raw", "")).strip() or None,
                "linea_raw": str(get(first_row, "linea_raw", "")).strip() or None,
                "peso_um": parse_number(get(first_row, "peso_um")),
                "is_profile": is_profile,
                "status": "active",
                "import_batch_id": batch_id,
            }

        # ── Build variant ──
        item_id = None
        for r in group:
            iid = get(r, "item_id")
            if iid:
                try:
                    item_id = int(iid)
                    break
                except (ValueError, TypeError):
                    pass

        variant_data[full_ref] = {
            "reference_siesa": full_ref,
            "_base_ref": base_ref,  # temp key for linking
            "item_id": item_id,
            "acabado_name": str(get(first_row, "acabado_desc_raw", "")).strip() or None,
            "acabado_code": str(get(first_row, "acabado_code_raw", "")).strip() or None,
            "aleacion": str(get(first_row, "aleacion_raw", "")).strip() or None,
            "longitud_m": length_m,
            "subcategoria_raw": subcategoria or None,
            "categoria_raw": str(get(first_row, "categoria_raw", "")).strip() or None,
            "sistema_raw": str(get(first_row, "sistema_raw", "")).strip() or None,
            "linea_raw": str(get(first_row, "linea_raw", "")).strip() or None,
            "status": "active",
            "flag_compra": any(parse_bool(get(r, "flag_compra")) for r in group),
            "flag_venta": any(parse_bool(get(r, "flag_venta")) for r in group),
            "flag_manufactura": any(parse_bool(get(r, "flag_manufactura")) for r in group),
            "flag_lote": any(parse_bool(get(r, "flag_lote")) for r in group),
            "flag_paquete": any(parse_bool(get(r, "flag_paquete")) for r in group),
            "posicion_arancelaria": str(get(first_row, "posicion_arancelaria", "")).strip() or None,
            "fecha_creacion_siesa": parse_date(get(first_row, "fecha_creacion_siesa")),
            "fecha_actualizacion_siesa": parse_date(get(first_row, "fecha_actualizacion_siesa")),
            "fecha_inactivacion": parse_date(get(first_row, "fecha_inactivacion")),
            "import_batch_id": batch_id,
        }

        # ── Build warehouse junction rows (one per bodega, dedup by key) ──
        for r in group:
            bodega = str(get(r, "bodega", "")).strip()
            if not bodega or bodega not in warehouse_map:
                continue
            wh_id = warehouse_map[bodega]
            dedup_key = (full_ref, wh_id)
            pw_data[dedup_key] = {
                "_full_ref": full_ref,
                "warehouse_id": wh_id,
                "costo_promedio": parse_number(get(r, "costo_promedio")),
                "precio_unitario": parse_number(get(r, "precio_unitario")),
                "costo_promedio_total": parse_number(get(r, "costo_promedio_total")),
                "margen_pct": parse_number(get(r, "margen_pct")),
                "unidad_precio": str(get(r, "unidad_precio", "")).strip() or None,
                "abc_rotacion_veces": str(get(r, "abc_rotacion_veces", "")).strip() or None,
                "abc_rotacion_costo": str(get(r, "abc_rotacion_costo", "")).strip() or None,
                "abc_rotacion_veces_bod": str(get(r, "abc_rotacion_veces_bod", "")).strip() or None,
                "abc_rotacion_costo_bod": str(get(r, "abc_rotacion_costo_bod", "")).strip() or None,
                "fecha_ultima_venta": parse_date(get(r, "fecha_ultima_venta")),
                "fecha_ultima_entrada": parse_date(get(r, "fecha_ultima_entrada")),
                "fecha_ultima_salida": parse_date(get(r, "fecha_ultima_salida")),
                "fecha_ultima_compra": parse_date(get(r, "fecha_ultima_compra")),
                "fecha_ultimo_conteo": parse_date(get(r, "fecha_ultimo_conteo")),
                "import_batch_id": batch_id,
            }

    pw_list = list(pw_data.values())
    print(f"  {len(product_data)} base products")
    print(f"  {len(variant_data)} variants")
    print(f"  {len(pw_list)} warehouse junction rows (deduped from {sum(len(g) for g in ref_groups.values())} raw rows)")

    # ═══ UPSERT PRODUCTS ═══
    print(f"\n  Upserting products...")
    products_list = list(product_data.values())
    for i in range(0, len(products_list), 500):
        batch = products_list[i:i + 500]
        supabase.table("products").upsert(batch, on_conflict="reference").execute()
        done = min(i + 500, len(products_list))
        if done % 500 == 0 or done == len(products_list):
            print(f"    products {done}/{len(products_list)}")

    # Fetch product IDs
    product_id_map = {}
    offset = 0
    while True:
        result = supabase.table("products").select("id, reference").range(offset, offset + 999).execute()
        for r in result.data:
            product_id_map[r["reference"]] = r["id"]
        if len(result.data) < 1000:
            break
        offset += 1000

    print(f"    {len(product_id_map)} product IDs fetched")

    # ═══ UPSERT VARIANTS ═══
    print(f"\n  Upserting variants...")
    variants_list = []
    for vd in variant_data.values():
        base_ref = vd.pop("_base_ref")
        vd["product_id"] = product_id_map.get(base_ref)
        if not vd["product_id"]:
            print(f"    WARNING: No product_id for base_ref={base_ref}")
            continue
        variants_list.append(vd)

    errors = 0
    for i in range(0, len(variants_list), 500):
        batch = variants_list[i:i + 500]
        try:
            supabase.table("product_variants").upsert(batch, on_conflict="reference_siesa").execute()
        except Exception as e:
            print(f"    ERROR batch {i}: {e}")
            errors += len(batch)
            continue
        done = min(i + 500, len(variants_list))
        if done % 500 == 0 or done == len(variants_list):
            print(f"    variants {done}/{len(variants_list)}")

    if errors:
        print(f"    {errors} variant errors")

    # Fetch variant IDs
    variant_id_map = {}
    offset = 0
    while True:
        result = supabase.table("product_variants").select("id, reference_siesa").range(offset, offset + 999).execute()
        for r in result.data:
            variant_id_map[r["reference_siesa"]] = r["id"]
        if len(result.data) < 1000:
            break
        offset += 1000

    print(f"    {len(variant_id_map)} variant IDs fetched")

    # ═══ UPSERT PRODUCT_WAREHOUSE ═══
    print(f"\n  Upserting product_warehouse junction...")
    pw_records = []
    for pw in pw_list:
        full_ref = pw.pop("_full_ref")
        vid = variant_id_map.get(full_ref)
        if not vid:
            continue
        pw["variant_id"] = vid
        pw_records.append(pw)

    pw_errors = 0
    for i in range(0, len(pw_records), 500):
        batch = pw_records[i:i + 500]
        try:
            supabase.table("product_warehouse").upsert(batch, on_conflict="variant_id,warehouse_id").execute()
        except Exception as e:
            print(f"    ERROR pw batch {i}: {e}")
            pw_errors += len(batch)
            continue
        done = min(i + 500, len(pw_records))
        if done % 2000 == 0 or done == len(pw_records):
            print(f"    product_warehouse {done}/{len(pw_records)}")

    print(f"    {len(pw_records) - pw_errors} junction rows upserted")
    if pw_errors:
        print(f"    {pw_errors} junction errors")

    return {
        "products": len(product_data),
        "variants": len(variants_list),
        "variant_errors": errors,
        "pw_rows": len(pw_records),
        "pw_errors": pw_errors,
        "skipped": skipped,
    }


# ═══════════════════════════════════════════════════════
# Step 4: Update classification SKU counts
# ═══════════════════════════════════════════════════════

def update_classification_counts():
    print(f"\nUpdating classification counts...")
    dim_field = {
        "categoria": "categoria_raw",
        "subcategoria": "subcategoria_raw",
        "sistema": "sistema_raw",
        "linea": "linea_raw",
    }
    # Count based on products table (base products)
    for dim, field in dim_field.items():
        try:
            supabase.rpc("exec_sql", {"query": f"""
                UPDATE classifications c
                SET sku_count = COALESCE(counts.cnt, 0)
                FROM (
                    SELECT {field} AS val, COUNT(*) AS cnt
                    FROM products WHERE {field} IS NOT NULL GROUP BY {field}
                ) counts
                WHERE c.dimension = '{dim}' AND c.original_value = counts.val
            """}).execute()
        except Exception:
            pass

    # acabado and aleacion counts from variants
    for dim, field in [("acabado", "acabado_name"), ("aleacion", "aleacion")]:
        try:
            supabase.rpc("exec_sql", {"query": f"""
                UPDATE classifications c
                SET sku_count = COALESCE(counts.cnt, 0)
                FROM (
                    SELECT {field} AS val, COUNT(*) AS cnt
                    FROM product_variants WHERE {field} IS NOT NULL GROUP BY {field}
                ) counts
                WHERE c.dimension = '{dim}' AND c.original_value = counts.val
            """}).execute()
        except Exception:
            pass

    print(f"  Counts updated")


# ═══════════════════════════════════════════════════════
# Step 5: Log import
# ═══════════════════════════════════════════════════════

def log_import(batch_id, file_name, file_size, rows_total, stats, status="completed"):
    supabase.table("import_history").insert({
        "id": batch_id,
        "file_type": "maestro",
        "file_name": file_name,
        "file_size_bytes": file_size,
        "rows_in_file": rows_total,
        "rows_new": stats["products"] + stats["variants"],
        "rows_updated": 0,
        "rows_skipped": stats["skipped"],
        "rows_error": stats["variant_errors"] + stats["pw_errors"],
        "status": status,
        "summary": {
            "products": stats["products"],
            "variants": stats["variants"],
            "warehouse_rows": stats["pw_rows"],
        },
        "completed_at": datetime.now().isoformat(),
    }).execute()
    print(f"\nImport logged: {batch_id}")


# ═══════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        filepath = r"C:\Users\Juan Felipe\OneDrive\Aragon Aluminio\Abastecimiento\COMMERCIAL SIESA\Maestro Commercial Items.xlsx"
    else:
        filepath = sys.argv[1]

    if not Path(filepath).exists():
        print(f"File not found: {filepath}")
        sys.exit(1)

    import uuid
    batch_id = str(uuid.uuid4())
    file_size = Path(filepath).stat().st_size
    file_name = Path(filepath).name

    print("=" * 60)
    print("  IMPORT MAESTRO DE ITEMS -> SUPABASE (Normalized)")
    print(f"  File: {file_name}")
    print(f"  Batch: {batch_id}")
    print("=" * 60)

    # 1. Read Excel
    headers, rows = read_excel(filepath)
    col_index = map_columns(headers)

    # 2. Classifications
    n_class = extract_and_upsert_classifications(rows, col_index)

    # 3. Warehouses
    warehouse_map = extract_and_upsert_warehouses(rows, col_index)

    # 4. Products + Variants + Warehouse junction
    stats = import_products_and_variants(rows, col_index, warehouse_map, batch_id)

    # 5. Update classification counts
    update_classification_counts()

    # 6. Log import
    log_import(batch_id, file_name, file_size, len(rows), stats)

    print(f"\n{'=' * 60}")
    print(f"  IMPORT COMPLETE")
    print(f"  Classifications: {n_class}")
    print(f"  Warehouses: {len(warehouse_map)}")
    print(f"  Products: {stats['products']}")
    print(f"  Variants: {stats['variants']}")
    print(f"  Warehouse rows: {stats['pw_rows']}")
    print(f"  Errors: {stats['variant_errors'] + stats['pw_errors']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
